import os
import json
import asyncio
import openpyxl
import requests
from telethon import TelegramClient, functions, types
from telethon.sessions import StringSession
from telegram import (
    Update, InlineKeyboardButton, InlineKeyboardMarkup, InputFile
)
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler, CallbackQueryHandler,
    ContextTypes, filters
)

# ====== CONFIGURATION =======
API_ID = int(os.getenv('API_ID'))
API_HASH = os.getenv('API_HASH')
BOT_TOKEN = os.getenv('BOT_TOKEN')
STRING_SESSION = os.getenv('STRING_SESSION')
PHONE_NUMBER = os.getenv('PHONE_NUMBER')

client = TelegramClient(StringSession(STRING_SESSION), API_ID, API_HASH)

CONFIG_FILE = 'config.json'

def load_config():
    if not os.path.exists(CONFIG_FILE):
        return {
            "tick_style": "✅",
            "last_summary": {"total_checked": 0, "registered_count": 0, "non_registered_count": 0},
            "non_registered_numbers": []
        }
    with open(CONFIG_FILE, 'r') as f:
        return json.load(f)

def save_config(config):
    with open(CONFIG_FILE, 'w') as f:
        json.dump(config, f, indent=4)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    config = load_config()
    summary = config['last_summary']
    tick = config['tick_style']

    keyboard = [
        [InlineKeyboardButton("📥 Upload Number List", callback_data='upload')],
        [InlineKeyboardButton("📄 Get Non-Registered TXT", callback_data='get_txt')],
        [InlineKeyboardButton(f"⚙️ Toggle {tick} / ✔️ Style", callback_data='toggle_style')],
        [InlineKeyboardButton("ℹ️ Help", callback_data='help')]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    message = (
        "🤖 Welcome to Telegram Number Checker Bot!\n\n"
        "Choose an option:\n\n"
        "[📥 Upload Number List]\n"
        "[📄 Get Non-Registered TXT]\n"
        "[⚙️ Toggle ✅ / ✔️ Style]\n"
        "[ℹ️ Help]\n\n"
        f"✅ Last Check Summary:\n"
        f"{tick} {summary['total_checked']} numbers checked\n"
        f"{tick} {summary['registered_count']} registered | "
        f"{tick} {summary['non_registered_count']} non-registered"
    )
    await update.message.reply_text(message, reply_markup=reply_markup)

async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    config = load_config()

    if query.data == 'upload':
        await query.message.reply_text("📁 Please send the XLSX file containing phone numbers (one number per row in first column).")
    elif query.data == 'get_txt':
        txt_path = 'non_registered_numbers.txt'
        with open(txt_path, 'w') as f:
            for num in config['non_registered_numbers']:
                f.write(f"{num}\n")
        await query.message.reply_document(document=InputFile(txt_path),
                                           caption="📄 Non-Registered Numbers TXT file.")
        os.remove(txt_path)
    elif query.data == 'toggle_style':
        config['tick_style'] = '✔️' if config['tick_style'] == '✅' else '✅'
        save_config(config)
        await query.message.reply_text(f"✅ Style toggled to: {config['tick_style']}")
    elif query.data == 'help':
        await query.message.reply_text(
            "📚 Usage Guide:\n"
            "📥 Upload a XLSX file (one phone number per row in the first column)\n"
            "📄 Download Non-Registered TXT file\n"
            "⚙️ Toggle between ✅ and ✔️ tick styles"
        )

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    file = await update.message.document.get_file()
    file_path = 'numbers.xlsx'
    await file.download_to_drive(file_path)

    await update.message.reply_text('⏳ Processing numbers, please wait...')

    config = load_config()
    tick = config['tick_style']
    non_registered_numbers = []
    total_checked = 0
    registered_count = 0

    await client.start(PHONE_NUMBER)
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    message_lines = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        number = row[0]

        if not number:
            continue  # Skip empty cells

        number = str(number).strip()
        if not number.startswith('+'):
            number = '+' + number

        try:
            result = await client(functions.contacts.ImportContactsRequest(
                contacts=[types.InputPhoneContact(client_id=0, phone=number, first_name='Test', last_name='User')]
            ))

            if result.users:
                registered_count += 1
                status = f"{tick} Registered"
                await client(functions.contacts.DeleteContactsRequest(id=[result.users[0]]))
            else:
                non_registered_numbers.append(number)
                status = "❌ Not Registered"

            message_lines.append(f"`{number}` - {status}")
            total_checked += 1
            await asyncio.sleep(1)

        except Exception:
            message_lines.append(f"`{number}` - ❌ Error")

    await client.disconnect()
    os.remove(file_path)

    config['last_summary'] = {
        'total_checked': total_checked,
        'registered_count': registered_count,
        'non_registered_count': len(non_registered_numbers)
    }
    config['non_registered_numbers'] = non_registered_numbers
    save_config(config)

    result_text = "\n".join(message_lines)
    await update.message.reply_markdown(f"✅ Check complete! Results:\n\n{result_text}")

    result_xlsx = 'checked_numbers.xlsx'
    output_wb = openpyxl.Workbook()
    output_sheet = output_wb.active
    output_sheet.append(['Phone Number', 'Status'])

    for line in message_lines:
        parts = line.split(' - ')
        output_sheet.append([parts[0].strip('`'), parts[1]])

    output_wb.save(result_xlsx)
    await update.message.reply_document(document=InputFile(result_xlsx),
                                        caption="📊 Result XLSX file.")
    os.remove(result_xlsx)

if __name__ == '__main__':
    import logging
    logging.basicConfig(level=logging.INFO)

    PORT = int(os.getenv('PORT', '8443'))
    HOST = '0.0.0.0'
    WEBHOOK_URL = f"https://your-render-app.onrender.com/{os.getenv('BOT_TOKEN')}"

    app = ApplicationBuilder().token(os.getenv('BOT_TOKEN')).build()
    app.add_handler(CommandHandler('start', start))
    app.add_handler(CallbackQueryHandler(button_handler))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))

    print(f"Running webhook on {HOST}:{PORT}")
    app.run_webhook(
        listen=HOST,
        port=PORT,
        WEBHOOK_URL = f"https://tg-bot-1jau.onrender.com/{BOT_TOKEN}"
    )
